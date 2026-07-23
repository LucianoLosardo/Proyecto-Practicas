import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

JSON_INPUT = "resultados_deteccion.json"
EXCEL_OUTPUT = "reporte_detecciones.xlsx"


def json_to_excel(json_path=JSON_INPUT, output_excel=EXCEL_OUTPUT):
    if not os.path.exists(json_path):
        print(f"Error: No se encontró el archivo '{json_path}'.")
        print("Asegúrate de ejecutar primero main.py para generar el JSON.")
        return

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()

    # Estilos de fuentes y colores
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="595959")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)

    fill_navy = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    fill_accent = PatternFill(
        start_color="2980B9", end_color="2980B9", fill_type="solid"
    )
    fill_zebra = PatternFill(
        start_color="F9FAFC", end_color="F9FAFC", fill_type="solid"
    )
    fill_card = PatternFill(
        start_color="EBF5FB", end_color="EBF5FB", fill_type="solid"
    )

    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    border_card = Border(
        left=Side(style="medium", color="1F4E79"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # -------------------------------------------------------------
    # HOJA 1: Resumen General
    # -------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Resumen General"
    ws_summary.views.sheetView[0].showGridLines = True

    # Encabezado principal
    ws_summary["A1"] = "Reporte de Detección de Fauna y Objetos - Wildcam"
    ws_summary["A1"].font = font_title
    meta = data.get("metadata", {})
    ws_summary["A2"] = (
        f"Modelo: {meta.get('model', 'N/A')} | Umbral: {meta.get('conf_threshold', 0.2)} | "
        f"Frame Skip: {meta.get('frame_skip', 10)} | "
        f"Tiempo de Ejecución: {meta.get('execution_time_formatted', 'N/A')}"
    )
    ws_summary["A2"].font = font_subtitle

    # Métricas clave (KPIs)
    videos = data.get("videos", [])
    total_vids = len(videos)
    vids_animals = sum(
        1
        for v in videos
        if "animal" in v.get("summary", {}).get("detected_classes", [])
    )
    vids_person = sum(
        1
        for v in videos
        if "person" in v.get("summary", {}).get("detected_classes", [])
    )
    vids_empty = sum(
        1 for v in videos if not v.get("summary", {}).get("detected_classes", [])
    )

    cards = [
        ("Total Videos", total_vids, "B4"),
        ("Videos con Fauna", vids_animals, "D4"),
        ("Presencia Humana", vids_person, "F4"),
        ("Videos Sin Detección", vids_empty, "H4"),
    ]

    for title, val, pos in cards:
        col = pos[0]
        row = int(pos[1])
        next_col = chr(ord(col) + 1)

        ws_summary.merge_cells(f"{pos}:{next_col}{row}")
        ws_summary.merge_cells(f"{col}{row+1}:{next_col}{row+1}")

        c_title = ws_summary[f"{col}{row}"]
        c_title.value = title
        c_title.font = Font(name="Segoe UI", size=9, bold=True, color="595959")
        c_title.alignment = align_center
        c_title.fill = fill_card

        c_val = ws_summary[f"{col}{row+1}"]
        c_val.value = val
        c_val.font = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
        c_val.alignment = align_center
        c_val.fill = fill_card

        for r in range(row, row + 2):
            for c in [ord(col) - 64, ord(next_col) - 64]:
                ws_summary.cell(row=r, column=c).border = border_card

    # Encabezados de la tabla de resumen
    headers_summary = [
        "Archivo de Video",
        "Detecciones Totales",
        "Clases Detectadas",
        "Máx. Confianza Animal",
        "Máx. Confianza Persona",
        "Estado",
    ]

    start_row = 8
    for col_idx, text in enumerate(headers_summary, 1):
        cell = ws_summary.cell(row=start_row, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_thin

    # Llenado de filas del resumen
    current_row = start_row + 1
    for idx, v in enumerate(videos, 1):
        summary = v.get("summary", {})
        det_classes = summary.get("detected_classes", [])
        max_conf = summary.get("max_confidence", {})

        conf_anim = max_conf.get("animal", 0.0)
        conf_pers = max_conf.get("person", 0.0)

        if "animal" in det_classes:
            status = "Fauna Detectada"
        elif "person" in det_classes:
            status = "Presencia Humana"
        elif "vehicle" in det_classes:
            status = "Vehículo Detectado"
        else:
            status = "Sin Detecciones"

        row_vals = [
            v.get("file", ""),
            v.get("filepath", ""),
            summary.get("total_detections_count", 0),
            ", ".join(det_classes) or "Ninguna",
            conf_anim if conf_anim else "-",
            conf_pers if conf_pers else "-",
            status,
        ]

        is_even = idx % 2 == 0
        row_fill = fill_zebra if is_even else PatternFill(fill_type=None)

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_summary.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_thin
            cell.fill = row_fill

            if col_idx in [1, 2, 4, 7]:
                cell.alignment = align_left
            elif col_idx == 3:
                cell.alignment = align_right
                cell.number_format = "#,##0"
            elif col_idx in [5, 6]:
                cell.alignment = align_right
                if isinstance(val, (int, float)):
                    cell.number_format = "0.0%"

        current_row += 1

    ws_summary.freeze_panes = f"A{start_row+1}"

    # Autoajuste de anchos
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)
    ws_summary.column_dimensions["B"].width = 35

    # -------------------------------------------------------------
    # HOJA 2: Detalle de Detecciones
    # -------------------------------------------------------------
    ws_detail = wb.create_sheet(title="Detalle Detecciones")
    ws_detail.views.sheetView[0].showGridLines = True

    headers_detail = [
        "Archivo Video",
        "Fotograma (Frame)",
        "Tiempo (Seg)",
        "Tiempo (mm:ss)",
        "Etiqueta / Clase",
        "Confianza (%)",
        "BBox Xmin",
        "BBox Ymin",
        "BBox Xmax",
        "BBox Ymax",
        "Ancho BBox",
        "Alto BBox",
    ]

    for col_idx, text in enumerate(headers_detail, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = align_center
        cell.border = border_thin

    detail_row = 2
    for v in videos:
        filename = v.get("file", "")
        for fd in v.get("frame_detections", []):
            frame = fd.get("frame", 0)
            t_sec = fd.get("time_sec", 0.0)

            mins = int(t_sec // 60)
            secs = int(t_sec % 60)
            time_str = f"{mins:02d}:{secs:02d}"

            for det in fd.get("detections", []):
                bbox = det.get("bbox", [0, 0, 0, 0])
                xmin, ymin, xmax, ymax = bbox[0], bbox[1], bbox[2], bbox[3]

                row_vals = [
                    filename,
                    frame,
                    t_sec,
                    time_str,
                    det.get("label", ""),
                    det.get("confidence", 0.0),
                    xmin,
                    ymin,
                    xmax,
                    ymax,
                    round(xmax - xmin, 1),
                    round(ymax - ymin, 1),
                ]

                is_even = detail_row % 2 == 0
                row_fill = fill_zebra if is_even else PatternFill(fill_type=None)

                for col_idx, val in enumerate(row_vals, 1):
                    cell = ws_detail.cell(
                        row=detail_row, column=col_idx, value=val
                    )
                    cell.font = font_data
                    cell.border = border_thin
                    cell.fill = row_fill

                    if col_idx in [1, 4, 5]:
                        cell.alignment = align_left
                    elif col_idx == 2:
                        cell.alignment = align_right
                        cell.number_format = "#,##0"
                    elif col_idx == 3:
                        cell.alignment = align_right
                        cell.number_format = "0.00"
                    elif col_idx == 6:
                        cell.alignment = align_right
                        cell.number_format = "0.0%"
                    else:
                        cell.alignment = align_right
                        cell.number_format = "0.0"

                detail_row += 1

    ws_detail.freeze_panes = "A2"
    ws_detail.auto_filter.ref = f"A1:L{detail_row-1}"

    for col in ws_detail.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_detail.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Guardar Excel
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(base_dir, output_excel)
    wb.save(output_path)
    print(f"Excel generado correctamente en: {output_path}")


if __name__ == "__main__":
    json_to_excel()